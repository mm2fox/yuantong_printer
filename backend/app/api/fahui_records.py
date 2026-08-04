from fastapi import APIRouter, Depends, HTTPException, status, Query, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_, and_, func
from typing import List, Optional
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from ..core.database import get_db
from ..models.user import User
from ..models.fahui_record import FahuiRecord
from ..models.fahui_user import FahuiUser
from ..models.system_log import SystemLog
from ..schemas.fahui_record import FahuiRecordCreate, FahuiRecordUpdate, FahuiRecordResponse
from .auth import get_current_user

router = APIRouter(prefix="/api/fahui-records", tags=["法会流水"])


def _record_to_dict(record, user=None):
    return {
        "id": record.id,
        "fahui_user_id": record.fahui_user_id,
        "fahui_id": record.fahui_id,
        "fahui_name": record.fahui_name,
        "座次": record.座次,
        "amount": record.amount,
        "paiwei_type": record.paiwei_type,
        "yanwang": record.yanwang,
        "xm1": record.xm1,
        "xm2": record.xm2,
        "xm3": record.xm3,
        "xm4": record.xm4,
        "xm5": record.xm5,
        "xm6": record.xm6,
        "xm7": record.xm7,
        "xm8": record.xm8,
        "xm9": record.xm9,
        "xm10": record.xm10,
        "xm": record.xm,
        "djdate": record.djdate,
        "经办人": record.经办人,
        "prt": record.prt,
        "remarks": record.remarks,
        "施主姓名": record.施主姓名 or (user.施主姓名 if user else None),
        "施主编号": record.施主编号 or (user.施主编号 if user else None),
        "电话": user.电话 if user else None,
        "功德主": user.功德主 if user else None
    }

@router.get("", response_model=List[FahuiRecordResponse])
async def get_fahui_records(
    fahui_name: Optional[str] = Query(None, description="法会名称"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = select(FahuiRecord).where(FahuiRecord.temple_id == current_user.temple_id)
    
    if fahui_name:
        query = query.where(FahuiRecord.fahui_name == fahui_name)
    
    result = await db.execute(query)
    records = result.scalars().all()
    return records

async def _fetch_fahui_query(
    db: AsyncSession,
    temple_id,
    fahui_name: Optional[str] = None,
    keyword: Optional[str] = None,
    shizhu_name: Optional[str] = None,
    shizhu_code: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    paiwei_type: Optional[str] = None,
    yanwang: Optional[int] = None,
    prt: Optional[int] = None,
    skip: int = 0,
    limit: Optional[int] = None
):
    base_where = [FahuiRecord.temple_id == temple_id]

    if fahui_name:
        base_where.append(FahuiRecord.fahui_name == fahui_name)
    if start_date:
        base_where.append(FahuiRecord.djdate >= start_date)
    if end_date:
        base_where.append(FahuiRecord.djdate <= end_date)
    if paiwei_type:
        base_where.append(FahuiRecord.paiwei_type == paiwei_type)
    if yanwang is not None:
        base_where.append(FahuiRecord.yanwang == yanwang)
    if prt is not None:
        base_where.append(FahuiRecord.prt == prt)
    if shizhu_name:
        base_where.append(FahuiRecord.施主姓名.contains(shizhu_name))
    if shizhu_code:
        base_where.append(FahuiRecord.施主编号.contains(shizhu_code))
    if keyword:
        base_where.append(or_(
            FahuiRecord.施主姓名.contains(keyword),
            FahuiRecord.施主编号.contains(keyword),
            FahuiRecord.fahui_name.contains(keyword),
            FahuiRecord.xm1.contains(keyword),
            FahuiRecord.xm2.contains(keyword),
            FahuiRecord.xm3.contains(keyword),
            FahuiRecord.xm4.contains(keyword),
            FahuiRecord.xm5.contains(keyword),
            FahuiRecord.xm6.contains(keyword),
            FahuiRecord.xm7.contains(keyword),
            FahuiRecord.xm8.contains(keyword),
            FahuiRecord.xm9.contains(keyword),
            FahuiRecord.xm10.contains(keyword),
            FahuiRecord.remarks.contains(keyword),
            FahuiRecord.经办人.contains(keyword),
            FahuiRecord.座次.contains(keyword),
        ))

    where_clause = and_(*base_where)

    count_query = select(func.count(FahuiRecord.id)).where(where_clause)
    count_result = await db.execute(count_query)
    total = count_result.scalar()

    sum_query = select(func.sum(FahuiRecord.amount)).where(where_clause)
    sum_result = await db.execute(sum_query)
    total_amount = sum_result.scalar() or 0

    query = select(FahuiRecord, FahuiUser).join(
        FahuiUser, FahuiRecord.fahui_user_id == FahuiUser.id, isouter=True
    ).where(where_clause).order_by(FahuiRecord.id.desc()).offset(skip)
    if limit is not None:
        query = query.limit(limit)

    result = await db.execute(query)
    rows = result.all()

    records = [_record_to_dict(record, user) for record, user in rows]
    return records, total, total_amount


@router.get("/query-by-fahui")
async def query_by_fahui(
    fahui_name: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    shizhu_name: Optional[str] = Query(None),
    shizhu_code: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    paiwei_type: Optional[str] = Query(None),
    yanwang: Optional[int] = Query(None),
    prt: Optional[int] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    records, total, total_amount = await _fetch_fahui_query(
        db, current_user.temple_id,
        fahui_name=fahui_name, keyword=keyword, shizhu_name=shizhu_name,
        shizhu_code=shizhu_code, start_date=start_date, end_date=end_date,
        paiwei_type=paiwei_type, yanwang=yanwang, prt=prt,
        skip=skip, limit=limit
    )
    return {
        "records": records,
        "total": total,
        "total_amount": total_amount
    }


async def _fetch_shizhu_query(
    db: AsyncSession,
    temple_id,
    shizhu_name: Optional[str] = None,
    shizhu_code: Optional[str] = None,
    phone: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    skip: int = 0,
    limit: Optional[int] = None
):
    base_where = [FahuiRecord.temple_id == temple_id]

    if shizhu_name:
        base_where.append(FahuiRecord.施主姓名.contains(shizhu_name))
    if shizhu_code:
        base_where.append(FahuiRecord.施主编号.contains(shizhu_code))
    if start_date:
        base_where.append(FahuiRecord.djdate >= start_date)
    if end_date:
        base_where.append(FahuiRecord.djdate <= end_date)

    where_no_phone = and_(*base_where)

    if phone:
        query_base = select(FahuiRecord, FahuiUser).join(
            FahuiUser, FahuiRecord.fahui_user_id == FahuiUser.id
        ).where(where_no_phone).where(FahuiUser.电话.contains(phone))
        count_base = select(func.count(FahuiRecord.id)).join(
            FahuiUser, FahuiRecord.fahui_user_id == FahuiUser.id
        ).where(where_no_phone).where(FahuiUser.电话.contains(phone))
        sum_base = select(func.sum(FahuiRecord.amount)).join(
            FahuiUser, FahuiRecord.fahui_user_id == FahuiUser.id
        ).where(where_no_phone).where(FahuiUser.电话.contains(phone))
    else:
        query_base = select(FahuiRecord, FahuiUser).join(
            FahuiUser, FahuiRecord.fahui_user_id == FahuiUser.id, isouter=True
        ).where(where_no_phone)
        count_base = select(func.count(FahuiRecord.id)).where(where_no_phone)
        sum_base = select(func.sum(FahuiRecord.amount)).where(where_no_phone)

    count_result = await db.execute(count_base)
    total = count_result.scalar()

    sum_result = await db.execute(sum_base)
    total_amount = sum_result.scalar() or 0

    query = query_base.order_by(FahuiRecord.id.desc()).offset(skip)
    if limit is not None:
        query = query.limit(limit)
    result = await db.execute(query)
    rows = result.all()

    records = [_record_to_dict(record, user) for record, user in rows]
    return records, total, total_amount


@router.get("/query-by-shizhu")
async def query_by_shizhu(
    shizhu_name: Optional[str] = Query(None),
    shizhu_code: Optional[str] = Query(None),
    phone: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    records, total, total_amount = await _fetch_shizhu_query(
        db, current_user.temple_id,
        shizhu_name=shizhu_name, shizhu_code=shizhu_code, phone=phone,
        start_date=start_date, end_date=end_date,
        skip=skip, limit=limit
    )
    return {
        "records": records,
        "total": total,
        "total_amount": total_amount
    }


def _build_excel_workbook(records, columns, sheet_name, total_amount=None):
    """构建带样式的 Excel 工作簿。columns: [(表头, 取值函数或字段名), ...]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell_align = Alignment(vertical="center", wrap_text=True)

    # 表头
    headers = [c[0] for c in columns]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # 数据行
    for item in records:
        row = []
        for _, getter in columns:
            if callable(getter):
                row.append(getter(item))
            else:
                row.append(item.get(getter))
        ws.append(row)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = cell_align
            cell.border = border

    # 合计行
    if total_amount is not None:
        total_row = ["合计"] + [None] * (len(headers) - 1)
        ws.append(total_row)
        total_cell = ws.cell(row=ws.max_row, column=1)
        total_cell.font = Font(bold=True)
        # 找到"金额"列填入合计
        for col_idx, (header, _) in enumerate(columns, start=1):
            if header == "金额":
                amt_cell = ws.cell(row=ws.max_row, column=col_idx)
                amt_cell.value = total_amount
                amt_cell.font = Font(bold=True)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.border = border

    # 列宽自适应（粗略）
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, min(ws.max_row + 1, 102)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    return wb


def _workbook_to_streaming_response(wb, filename):
    from urllib.parse import quote
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # HTTP 头只能用 latin-1 编码,中文文件名需用 RFC 5987 的 filename* 格式
    encoded = quote(filename)
    headers = {
        "Content-Disposition": f"attachment; filename=\"export.xlsx\"; filename*=UTF-8''{encoded}"
    }
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


def _fahui_export_columns():
    return [
        ("施主编号", "施主编号"),
        ("施主姓名", "施主姓名"),
        ("法会名称", "fahui_name"),
        ("牌位类型", "paiwei_type"),
        ("金额", "amount"),
        ("类型", lambda r: "延生" if r.get("yanwang") == 0 else "往生"),
        ("登记日期", "djdate"),
        ("经办人", "经办人"),
        ("打印状态", lambda r: "已打印" if r.get("prt") == 1 else "未打印"),
        ("姓名1", "xm1"),
        ("姓名2", "xm2"),
        ("姓名3", "xm3"),
        ("姓名4", "xm4"),
        ("姓名5", "xm5"),
        ("备注", "remarks"),
    ]


def _shizhu_export_columns():
    return [
        ("施主编号", "施主编号"),
        ("施主姓名", "施主姓名"),
        ("电话", "电话"),
        ("法会名称", "fahui_name"),
        ("牌位类型", "paiwei_type"),
        ("金额", "amount"),
        ("类型", lambda r: "延生" if r.get("yanwang") == 0 else "往生"),
        ("登记日期", "djdate"),
        ("经办人", "经办人"),
        ("打印状态", lambda r: "已打印" if r.get("prt") == 1 else "未打印"),
        ("姓名1", "xm1"),
        ("姓名2", "xm2"),
        ("姓名3", "xm3"),
        ("姓名4", "xm4"),
        ("姓名5", "xm5"),
        ("备注", "remarks"),
    ]


@router.get("/export-by-fahui")
async def export_by_fahui(
    fahui_name: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    shizhu_name: Optional[str] = Query(None),
    shizhu_code: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    paiwei_type: Optional[str] = Query(None),
    yanwang: Optional[int] = Query(None),
    prt: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    records, total, total_amount = await _fetch_fahui_query(
        db, current_user.temple_id,
        fahui_name=fahui_name, keyword=keyword, shizhu_name=shizhu_name,
        shizhu_code=shizhu_code, start_date=start_date, end_date=end_date,
        paiwei_type=paiwei_type, yanwang=yanwang, prt=prt,
        skip=0, limit=None
    )
    wb = _build_excel_workbook(records, _fahui_export_columns(), "法会记录", total_amount=total_amount)
    today = datetime.now().strftime("%Y%m%d")
    return _workbook_to_streaming_response(wb, f"法会记录查询_{today}.xlsx")


@router.get("/export-by-shizhu")
async def export_by_shizhu(
    shizhu_name: Optional[str] = Query(None),
    shizhu_code: Optional[str] = Query(None),
    phone: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    records, total, total_amount = await _fetch_shizhu_query(
        db, current_user.temple_id,
        shizhu_name=shizhu_name, shizhu_code=shizhu_code, phone=phone,
        start_date=start_date, end_date=end_date,
        skip=0, limit=None
    )
    wb = _build_excel_workbook(records, _shizhu_export_columns(), "施主法会记录", total_amount=total_amount)
    today = datetime.now().strftime("%Y%m%d")
    return _workbook_to_streaming_response(wb, f"施主查询_{today}.xlsx")


@router.post("/batch")
async def batch_create_fahui_records(
    payload: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    records_data = payload.get("records", [])
    if not records_data:
        raise HTTPException(status_code=400, detail="记录列表不能为空")

    fahui_id = payload.get("fahui_id")
    fahui_name = payload.get("fahui_name")
    djdate = payload.get("djdate") or datetime.now().strftime("%Y-%m-%d")

    if fahui_id:
        max_result = await db.execute(
            select(FahuiRecord.座次).where(FahuiRecord.fahui_id == fahui_id).order_by(FahuiRecord.座次.desc()).limit(1)
        )
        max_zuoci = max_result.scalar_one_or_none()
        next_zuoci = (int(max_zuoci) + 1) if max_zuoci else 1
    else:
        next_zuoci = 1

    created_records = []
    for item in records_data:
        record = FahuiRecord(
            fahui_id=fahui_id,
            fahui_name=fahui_name,
            fahui_user_id=item.get("fahui_user_id"),
            yanwang=item.get("yanwang", 0),
            paiwei_type=item.get("paiwei_type", "中牌"),
            amount=item.get("amount", 0),
            xm1=item.get("xm1"),
            xm2=item.get("xm2"),
            xm3=item.get("xm3"),
            xm4=item.get("xm4"),
            xm5=item.get("xm5"),
            xm6=item.get("xm6"),
            xm7=item.get("xm7"),
            xm8=item.get("xm8"),
            xm9=item.get("xm9"),
            xm10=item.get("xm10"),
            xm=item.get("xm"),
            djdate=djdate,
            prt=0,
            remarks=item.get("remarks"),
            座次=str(next_zuoci),
            经办人=current_user.username,
            施主姓名=item.get("施主姓名"),
            施主编号=item.get("施主编号"),
        )
        if current_user.temple_id:
            record.temple_id = current_user.temple_id
        db.add(record)
        created_records.append(record)
        next_zuoci += 1

    await db.commit()
    for r in created_records:
        await db.refresh(r)

    log = SystemLog(
        用户名=current_user.username,
        操作类型="批量新增",
        操作内容=f"批量创建法会登记：{fahui_name} - 共{len(created_records)}条",
        created_at=datetime.utcnow()
    )
    db.add(log)
    await db.commit()

    return {"message": f"成功创建{len(created_records)}条记录", "count": len(created_records)}

@router.get("/{record_id}", response_model=FahuiRecordResponse)
async def get_fahui_record(
    record_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    result = await db.execute(select(FahuiRecord).where(FahuiRecord.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    return record

@router.post("", response_model=FahuiRecordResponse)
async def create_fahui_record(
    record_data: FahuiRecordCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if record_data.fahui_id:
        max_result = await db.execute(
            select(FahuiRecord.座次).where(FahuiRecord.fahui_id == record_data.fahui_id).order_by(FahuiRecord.座次.desc()).limit(1)
        )
        max_zuoci = max_result.scalar_one_or_none()
        next_zuoci = (int(max_zuoci) + 1) if max_zuoci else 1
    else:
        next_zuoci = 1
    
    record = FahuiRecord(**record_data.dict())
    record.座次 = str(next_zuoci)
    record.经办人 = current_user.username
    
    if record_data.fahui_user_id:
        user_result = await db.execute(select(FahuiUser).where(FahuiUser.id == record_data.fahui_user_id))
        shizhu = user_result.scalar_one_or_none()
        if shizhu:
            record.施主姓名 = shizhu.施主姓名
            record.施主编号 = shizhu.施主编号
    
    if record_data.djdate:
        record.djdate = record_data.djdate
    else:
        record.djdate = datetime.now().strftime("%Y-%m-%d")
    if current_user.temple_id:
        record.temple_id = current_user.temple_id
    
    db.add(record)
    await db.commit()
    await db.refresh(record)
    
    log = SystemLog(
        用户名=current_user.username,
        操作类型="新增",
        操作内容=f"法会登记：{record.fahui_name} - {record.施主姓名}（{record.amount}元）",
        created_at=datetime.utcnow()
    )
    db.add(log)
    await db.commit()
    
    return record

@router.put("/{record_id}", response_model=FahuiRecordResponse)
async def update_fahui_record(
    record_id: int,
    record_data: FahuiRecordUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    result = await db.execute(select(FahuiRecord).where(FahuiRecord.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    
    update_data = record_data.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(record, key, value)

    # 编辑记录数据时自动更新经办人为当前操作用户（仅更新打印状态时不改经办人）
    if any(key != 'prt' for key in update_data.keys()):
        record.经办人 = current_user.username

    await db.commit()
    await db.refresh(record)

    log = SystemLog(
        用户名=current_user.username,
        操作类型="修改",
        操作内容=f"修改法会登记：{record.fahui_name} - {record.施主姓名}",
        created_at=datetime.utcnow()
    )
    db.add(log)
    await db.commit()
    
    return record

@router.delete("/{record_id}")
async def delete_fahui_record(
    record_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    result = await db.execute(select(FahuiRecord).where(FahuiRecord.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    
    fahui_name = record.fahui_name
    shizhu_name = record.施主姓名
    
    await db.delete(record)
    await db.commit()
    
    log = SystemLog(
        用户名=current_user.username,
        操作类型="删除",
        操作内容=f"删除法会登记：{fahui_name} - {shizhu_name}",
        created_at=datetime.utcnow()
    )
    db.add(log)
    await db.commit()
    
    return {"message": "删除成功"}
